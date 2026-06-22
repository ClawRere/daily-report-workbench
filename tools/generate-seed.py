from __future__ import annotations

import json
import re
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

from bs4 import BeautifulSoup
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "data" / "seed-data.js"
EXCEL_PATH = Path(r"C:\Users\94522\Desktop\2026年工作内容.xlsx")
HTML_PATH = Path(r"C:\Users\94522\.qclaw\workspace-agent-6d11c186\work-report.html")

PROJECT_RULES = [
    ("航空中心", ["航空中心", "航空航天", "航天中心", "航空项目"]),
    ("南航", ["南航"]),
    ("安利", ["安利"]),
    ("富力海洋世界", ["富力海洋世界", "海洋欢乐世界", "欢乐海洋世界"]),
    ("左庭右院", ["左庭右院"]),
    ("奈尔宝", ["奈尔宝"]),
    ("拾味馆", ["拾味馆"]),
    ("斐乐", ["斐乐", "fila", "Fila"]),
    ("德克士", ["德克士", "德科"]),
    ("华住", ["华住"]),
    ("国寿", ["国寿"]),
    ("友邦", ["友邦"]),
    ("亚朵", ["亚朵"]),
    ("泉膳", ["泉膳"]),
    ("POPEYES", ["POPEYES", "popeyes"]),
    ("起亚", ["起亚"]),
    ("东风汽车", ["东风汽车"]),
    ("欧派", ["欧派"]),
    ("亚玛芬", ["亚玛芬"]),
    ("工单", ["工单"]),
    ("玛氏", ["玛氏"]),
    ("天财", ["天财"]),
    ("华润", ["华润"]),
    ("问卷网", ["问卷网"]),
    ("历峰", ["历峰"]),
    ("wagas", ["wagas"]),
    ("海康威视", ["海康威视"]),
    ("美丽田园", ["美丽田园"]),
    ("AI虚拟人", ["AI虚拟人", "虚拟人"]),
    ("手机AI", ["手机AI"]),
    ("山东大学", ["山东大学"]),
    ("亿美", ["亿美"]),
]


@dataclass
class Task:
    id: str
    date: str
    project: str
    title: str
    category: str
    priority: str
    status: str
    plan: str
    notes: str
    source: str


def normalize_project(text: str) -> str:
    for project, keywords in PROJECT_RULES:
        if any(keyword in text for keyword in keywords):
            return project
    match = re.search(r"([\u4e00-\u9fa5A-Za-z0-9]+项目)", text)
    if match:
        return match.group(1)
    return "日常事务"


def normalize_category(text: str) -> str:
    if re.search(r"请假|假期", text):
        return "休假"
    if re.search(r"面试|实习生|培训|入职", text):
        return "招聘培训"
    if re.search(r"发票|报销|预付款|供应商|总结|汇报|交接|出差|计划书", text):
        return "行政协同"
    if re.search(r"评审|沟通|对接|确认|会", text):
        return "需求沟通"
    if re.search(r"测试|提测|上线|验证|排查", text):
        return "测试上线"
    if re.search(r"数据|看板|BI|采集|报表|仪表盘|拼接", text):
        return "数据分析"
    if re.search(r"PRD|原型|文档|流程图|设计|手册|框架|规格", text):
        return "产品设计"
    if re.search(r"AI|虚拟人", text):
        return "AI探索"
    return "综合事务"


def normalize_priority(text: str) -> str:
    if re.search(r"漏洞|异常|紧急|问题处理|上线|评审|现场", text):
        return "高"
    if re.search(r"PRD|需求|优化|测试|更新|采集|设计|评审", text):
        return "中"
    return "普通"


def excel_tasks() -> Iterable[Task]:
    if not EXCEL_PATH.exists():
        return []
    workbook = load_workbook(EXCEL_PATH, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    output = []
    counter = 0
    for raw_date, raw_text in sheet.iter_rows(values_only=True):
        if not raw_date or not raw_text:
            continue
        if isinstance(raw_date, datetime):
            date_str = raw_date.strftime("%Y-%m-%d")
        else:
            continue
        parts = [part.strip("；; ") for part in re.split(r"[；;]+", str(raw_text)) if part and part.strip("；; ")]
        for part in parts:
            counter += 1
            output.append(
                Task(
                    id=f"excel-{counter}",
                    date=date_str,
                    project=normalize_project(part),
                    title=part,
                    category=normalize_category(part),
                    priority=normalize_priority(part),
                    status="planned" if re.search(r"请假|假期", part) else "done",
                    plan="",
                    notes="来自 2026 年工作内容表",
                    source="excel",
                )
            )
    return output


def html_tasks() -> Iterable[Task]:
    if not HTML_PATH.exists():
        return []
    soup = BeautifulSoup(HTML_PATH.read_text(encoding="utf-8-sig"), "html.parser")
    output = []
    counter = 0
    for day_card in soup.select(".day-card"):
        day_id = day_card.get("id", "")
        if not day_id.startswith("day-"):
            continue
        date_str = day_id.replace("day-", "")
        for task_item in day_card.select(".task-item"):
            title_el = task_item.select_one(".task-text")
            if not title_el:
                continue
            title = title_el.get_text(strip=True)
            if not title:
                continue
            tags = [tag.get_text(strip=True) for tag in task_item.select(".tag")]
            status = "done"
            marker = task_item.select_one(".tc")
            marker_classes = marker.get("class", []) if marker else []
            if "tc-pending" in marker_classes:
                status = "pending"
            if "tc-delay" in marker_classes:
                status = "delayed"

            project = normalize_project(title)
            explicit_project_tags = [
                tag
                for tag in tags
                if tag not in {"已完成", "待办", "个人", "数据", "评审", "行政", "招聘", "分享", "其他", "多次延期"}
            ]
            if explicit_project_tags:
                project = explicit_project_tags[0]

            notes = "；".join(tag for tag in tags if tag not in {project, "已完成"})
            counter += 1
            output.append(
                Task(
                    id=f"html-{counter}",
                    date=date_str,
                    project=project,
                    title=title,
                    category=normalize_category(title + " " + " ".join(tags)),
                    priority=normalize_priority(title),
                    status=status,
                    plan="",
                    notes=notes,
                    source="html",
                )
            )
    return output


def dedupe(tasks: Iterable[Task]) -> list[Task]:
    seen = set()
    output = []
    for task in tasks:
        key = (task.date, task.project, task.title, task.status)
        if key in seen:
            continue
        seen.add(key)
        output.append(task)
    output.sort(key=lambda item: (item.date, item.title), reverse=True)
    return output


def main() -> None:
    tasks = dedupe([*excel_tasks(), *html_tasks()])
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    # Keep the seed file encoding-safe for static hosting by writing escaped Unicode.
    payload = json.dumps([asdict(task) for task in tasks], ensure_ascii=True, indent=2)
    OUTPUT.write_text(f"window.SEED_DATA = {payload};\n", encoding="utf-8")
    print(f"seed tasks: {len(tasks)}")
    print(f"output: {OUTPUT}")


if __name__ == "__main__":
    main()
